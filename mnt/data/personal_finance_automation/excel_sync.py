import os
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from database import connect

BASE = Path(__file__).resolve().parent
EXCEL_PATH = Path(os.getenv("EXCEL_PATH", BASE / "Personal_Financial_Model_V2.xlsx")).resolve()


def clear_data_rows(ws, start_row, max_col):
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.value = None


def write_sheet_table(ws, start_row, rows, max_col):
    for i, row in enumerate(rows, start_row):
        for j, value in enumerate(row, 1):
            ws.cell(i, j).value = value


def sync_workbook():
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {EXCEL_PATH}")

    wb = load_workbook(EXCEL_PATH)
    with connect() as con:
        accounts = con.execute("""
            SELECT account_id,institution,account_type,account_name,current_balance,currency,last_sync,external_account_id,''
            FROM accounts ORDER BY account_id
        """).fetchall()
        transactions = con.execute("""
            SELECT date,account_id,transaction_id,merchant,transaction_type,category,subcategory,amount,
                   substr(date,1,7),recurring,source,
                   CASE WHEN category='Uncategorized' OR account_id IS NULL THEN 'Review' ELSE 'OK' END,
                   description
            FROM transactions ORDER BY date DESC, transaction_id DESC
        """).fetchall()
        rules = con.execute("""
            SELECT 'R'||printf('%03d',id),CASE WHEN active=1 THEN 'Yes' ELSE 'No' END,priority,'Contains',pattern,category,subcategory,transaction_type,CASE WHEN recurring=1 THEN 'Yes' ELSE 'No' END,''
            FROM category_rules ORDER BY priority,id
        """).fetchall()
        raw = con.execute("""
            SELECT transaction_id,date,account_id,description,CASE WHEN transaction_type='Expense' THEN -amount ELSE amount END,currency,
                   merchant,NULL,pending,transaction_id,source,created_at,fingerprint,''
            FROM transactions ORDER BY date DESC, transaction_id DESC
        """).fetchall()

    # Preserve workbook formatting and structure; replace data only.
    ws = wb["Accounts"]
    clear_data_rows(ws,4,9); write_sheet_table(ws,4,accounts,9)
    for r in range(4,4+len(accounts)): ws.cell(r,5).number_format='$#,##0.00'

    ws = wb["Raw Import"]
    clear_data_rows(ws,5,14); write_sheet_table(ws,5,raw,14)
    for r in range(5,5+len(raw)):
        ws.cell(r,2).number_format='mm/dd/yyyy'; ws.cell(r,5).number_format='$#,##0.00'; ws.cell(r,12).number_format='mm/dd/yyyy hh:mm'

    ws = wb["Transactions"]
    clear_data_rows(ws,5,13); write_sheet_table(ws,5,transactions,13)
    for r in range(5,5+len(transactions)):
        ws.cell(r,1).number_format='mm/dd/yyyy'; ws.cell(r,8).number_format='$#,##0.00'

    ws = wb["Category Rules"]
    clear_data_rows(ws,4,10); write_sheet_table(ws,4,rules,10)

    wb.save(EXCEL_PATH)
    return str(EXCEL_PATH)

if __name__ == '__main__':
    print(sync_workbook())
